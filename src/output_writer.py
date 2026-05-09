from __future__ import annotations  # 支持新式类型注解

from pathlib import Path  # 输出路径处理
from typing import Any, Dict, List  # 类型注解

import pandas as pd  # 使用 DataFrame 写 Excel
from openpyxl import load_workbook  # 写完 Excel 后再打开做样式
from openpyxl.styles import Alignment, Font, PatternFill  # Excel 样式：对齐、字体、填充
from openpyxl.utils import get_column_letter  # 列号转 Excel 字母列


## 输出写入器：负责生成提取结果、模型比对、最终复核等 Excel 文件
class OutputWriter:
    ## 初始化：缓存字段清单和主 sheet 名，确保所有 Excel 使用同一模板
    def __init__(self, config: Dict[str, Any]):
        self.config = config  # 保存全局配置
        self.fields = config.get("extraction", {}).get("fields", [])  # 提取结果字段清单
        self.sheet_name = config.get("extraction", {}).get("sheet_name", "规范解读")  # 主结果 sheet 名

    ## 写出所有输出：每个输入文件会调用一次该方法
    def write_all(
        self,
        run_dir: Path,  # 当前文件独立输出目录
        input_file: Path,  # 当前输入文件路径
        document: Dict[str, Any],  # 解析后的文档结构，预留后续输出使用
        chunks: List[Any],  # RAG chunk 列表
        model_outputs: Dict[str, Any],  # 模型 A/B 输出
        comparison: List[Dict[str, Any]],  # 字段比对明细
        review_rows: List[Dict[str, Any]],  # 人工校验行
        final_rows: List[Dict[str, Any]],  # 最终结果初稿行
    ) -> None:
        if self.config.get("output", {}).get("save_individual_model_excels", True):  # 是否输出模型单独提取结果
            self._write_extraction_excel(run_dir / "extracted_model_a.xlsx", model_outputs.get("model_a", {}).get("rows", []))  # 写模型 A 结果
            self._write_extraction_excel(run_dir / "extracted_model_b.xlsx", model_outputs.get("model_b", {}).get("rows", []))  # 写模型 B 结果

        self._write_comparison_excel(run_dir / "comparison_result.xlsx", input_file, model_outputs, comparison, review_rows, chunks)  # 写模型比对文件
        self._write_final_excel(run_dir / self.config.get("output", {}).get("excel_name", "final_review_result.xlsx"), final_rows, review_rows, comparison, model_outputs, chunks)  # 写最终复核文件

    ## 写单个模型的提取结果 Excel：格式对齐“提取结果”模板
    def _write_extraction_excel(self, path: Path, rows: List[Dict[str, Any]]) -> None:
        rows = rows or [{field: "" for field in self.fields}]  # 如果模型没有有效行，至少输出一行空模板
        df = pd.DataFrame(rows, columns=self.fields)  # 严格按字段顺序构建 DataFrame
        with pd.ExcelWriter(path, engine="openpyxl") as writer:  # 使用 openpyxl 引擎便于后续样式处理
            df.to_excel(writer, sheet_name=self.sheet_name, index=False, startrow=1)  # 从第二行开始写表头，第一行留标题
        self._style_excel(path, title_by_sheet={self.sheet_name: self.sheet_name})  # 统一设置标题、表头、列宽

    ## 写模型比对 Excel：包含比对概览、字段比对、人工校验、RAG 召回
    def _write_comparison_excel(
        self,
        path: Path,  # 输出文件路径
        input_file: Path,  # 输入文件路径
        model_outputs: Dict[str, Any],  # 模型输出集合
        comparison: List[Dict[str, Any]],  # 比对明细
        review_rows: List[Dict[str, Any]],  # 人工校验行
        chunks: List[Any],  # RAG chunks
    ) -> None:
        model_a_label = model_outputs.get("model_a", {}).get("label", self.config.get("models", {}).get("model_a", {}).get("label", "模型A"))  # 模型 A 展示列名
        model_b_label = model_outputs.get("model_b", {}).get("label", self.config.get("models", {}).get("model_b", {}).get("label", "模型B"))  # 模型 B 展示列名

        ## 比对概览 sheet：记录文件、模型、阈值、需要复核数量等运行摘要
        overview = [
            {"项目": "输入文件", "值": str(input_file)},  # 当前处理文件
            {"项目": "左侧模型标签", "值": model_a_label},  # 左侧模型标签
            {"项目": "左侧模型名称", "值": self.config.get("models", {}).get("model_a", {}).get("model_name", "")},  # 左侧模型 API 名
            {"项目": "右侧模型标签", "值": model_b_label},  # 右侧模型标签
            {"项目": "右侧模型名称", "值": self.config.get("models", {}).get("model_b", {}).get("model_name", "")},  # 右侧模型 API 名
            {"项目": "相似度阈值", "值": self.config.get("rag", {}).get("similarity_threshold", 0.8)},  # 人工复核阈值
            {"项目": "需要人工复核字段数", "值": len(review_rows)},  # 进入人工校验的记录数
        ]

        ## 字段比对 sheet：把内部 model_a_value/model_b_value 映射成真实模型标签列名
        comparison_rows = [
            {
                "字段": item.get("字段", ""),  # 字段名
                model_a_label: item.get("model_a_value", ""),  # 左侧模型值
                model_b_label: item.get("model_b_value", ""),  # 右侧模型值
                "比对结论": item.get("比对结论", ""),  # 完全一致/基本一致/明显差异
                "相似度": f"{float(item.get('相似度', 0)):.3f}",  # 三位小数
                "说明": item.get("说明", ""),  # 阈值说明
            }
            for item in comparison  # 遍历所有比对明细
        ]

        review_df = pd.DataFrame(review_rows, columns=self._review_columns(model_a_label, model_b_label))  # 即使没有复核行，也保留固定列头

        ## 写入多个 sheet：startrow=1 让第一行作为合并标题行
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pd.DataFrame(overview).to_excel(writer, sheet_name="比对概览", index=False, startrow=1)  # 概览 sheet
            pd.DataFrame(comparison_rows).to_excel(writer, sheet_name="字段比对", index=False, startrow=1)  # 字段比对 sheet
            review_df.to_excel(writer, sheet_name="人工校验", index=False, startrow=1)  # 人工校验 sheet
            pd.DataFrame([chunk.to_dict() for chunk in chunks]).to_excel(writer, sheet_name="RAG召回", index=False, startrow=1)  # chunk 明细 sheet

        self._style_excel(
            path,
            title_by_sheet={
                "比对概览": "比对概览",  # 概览标题
                "字段比对": "字段比对",  # 字段比对标题
                "人工校验": "人工校验",  # 人工校验标题
                "RAG召回": "RAG召回",  # RAG 标题
            },
        )

    ## 写最终复核 Excel：这是后续人工修改和交付的主文件
    def _write_final_excel(
        self,
        path: Path,  # 输出文件路径
        final_rows: List[Dict[str, Any]],  # 最终结果初稿
        review_rows: List[Dict[str, Any]],  # 人工校验数据
        comparison: List[Dict[str, Any]],  # 比对明细
        model_outputs: Dict[str, Any],  # 模型输出
        chunks: List[Any],  # RAG chunk
    ) -> None:
        model_a_label = self.config.get("models", {}).get("model_a", {}).get("label", "模型A")  # 模型 A 标签
        model_b_label = self.config.get("models", {}).get("model_b", {}).get("label", "模型B")  # 模型 B 标签
        review_df = pd.DataFrame(review_rows, columns=self._review_columns(model_a_label, model_b_label))  # 固定人工校验列

        ## 写最终结果和所有可追溯辅助 sheet
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pd.DataFrame(final_rows, columns=self.fields).to_excel(writer, sheet_name=self.sheet_name, index=False, startrow=1)  # 主结果 sheet
            review_df.to_excel(writer, sheet_name="人工校验", index=False, startrow=1)  # 人工校验 sheet
            pd.DataFrame(comparison).to_excel(writer, sheet_name="字段比对明细", index=False, startrow=1)  # 内部完整比对明细
            pd.DataFrame(model_outputs.get("model_a", {}).get("rows", []), columns=self.fields).to_excel(writer, sheet_name="模型A提取结果", index=False, startrow=1)  # 模型 A 原始结构化结果
            pd.DataFrame(model_outputs.get("model_b", {}).get("rows", []), columns=self.fields).to_excel(writer, sheet_name="模型B提取结果", index=False, startrow=1)  # 模型 B 原始结构化结果
            pd.DataFrame([chunk.to_dict() for chunk in chunks]).to_excel(writer, sheet_name="RAG召回", index=False, startrow=1)  # RAG chunk 列表
            pd.DataFrame(flatten_config(self.config)).to_excel(writer, sheet_name="run_config", index=False, startrow=1)  # 运行配置快照

        self._style_excel(
            path,
            title_by_sheet={
                self.sheet_name: self.sheet_name,  # 主结果标题
                "人工校验": "人工校验",  # 人工校验标题
                "字段比对明细": "字段比对明细",  # 明细标题
                "模型A提取结果": "模型A提取结果",  # 模型 A 标题
                "模型B提取结果": "模型B提取结果",  # 模型 B 标题
                "RAG召回": "RAG召回",  # RAG 标题
                "run_config": "run_config",  # 配置标题
            },
        )

    ## 统一 Excel 样式：合并标题行、冻结窗格、自动列宽、表头填充
    def _style_excel(self, path: Path, title_by_sheet: Dict[str, str]) -> None:
        workbook = load_workbook(path)  # 打开刚写出的工作簿
        fill = PatternFill("solid", fgColor="D9EAF7")  # 表头浅蓝填充
        title_fill = PatternFill("solid", fgColor="1F4E78")  # 标题深蓝填充
        title_font = Font(color="FFFFFF", bold=True, size=12)  # 标题白色粗体
        header_font = Font(bold=True)  # 表头粗体

        ## 遍历需要设置标题的 sheet：不存在的 sheet 跳过，增强兼容性
        for sheet_name, title in title_by_sheet.items():
            if sheet_name not in workbook.sheetnames:  # 防御性检查 sheet 是否存在
                continue  # 不存在则跳过
            sheet = workbook[sheet_name]  # 取得工作表
            sheet.cell(row=1, column=1, value=title)  # 第一行第一列写标题
            sheet.cell(row=1, column=1).fill = title_fill  # 设置标题背景色
            sheet.cell(row=1, column=1).font = title_font  # 设置标题字体
            if sheet.max_column > 1:  # 多列时才合并标题行
                sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)  # 合并第一行

            for cell in sheet[2]:  # 第二行是 pandas 写出的表头
                cell.fill = fill  # 设置表头背景
                cell.font = header_font  # 设置表头字体
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 表头居中换行

            for row in sheet.iter_rows(min_row=3):  # 数据区从第三行开始
                for cell in row:  # 遍历每个数据单元格
                    cell.alignment = Alignment(vertical="top", wrap_text=True)  # 顶部对齐并自动换行

            for column_cells in sheet.columns:  # 自动估算每一列宽度
                column_letter = get_column_letter(column_cells[0].column)  # 获取 Excel 列字母
                max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])  # 只扫描前 200 行，避免大文件太慢
                sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 45)  # 列宽限制在 12-45
            sheet.freeze_panes = "A3"  # 冻结标题行和表头行

        workbook.save(path)  # 保存样式修改

    ## 人工校验固定列：即使没有低相似度记录，也要输出完整列头
    def _review_columns(self, model_a_label: str, model_b_label: str) -> List[str]:
        return [
            "区域",  # 比对区域
            "位置",  # 行位置
            "字段",  # 字段名
            model_a_label,  # 模型 A 列
            model_b_label,  # 模型 B 列
            "比对结论",  # 比对结论
            "相似度",  # 相似度
            "说明",  # 说明
            "RAG召回内容",  # 原文召回证据
            "最终建议值",  # 程序建议值
            "人工复核结果",  # 人工填写结果
            "复核备注",  # 人工备注
        ]


## 将嵌套 config 展平成两列表格，写入 run_config sheet
def flatten_config(config: Dict[str, Any], prefix: str = "") -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []  # 保存展平后的配置项
    for key, value in config.items():  # 遍历当前层级配置
        path = f"{prefix}.{key}" if prefix else str(key)  # 拼接层级路径
        if isinstance(value, dict):  # 如果值仍然是字典
            rows.extend(flatten_config(value, path))  # 递归展平子配置
        else:
            rows.append({"配置项": path, "值": value})  # 叶子节点写成一行
    return rows  # 返回展平配置表
